"""
ET Brands Dashboard — Script de actualización automática
Corre en GitHub Actions. Lee las 8 fuentes de Google Drive
y actualiza index.html directamente.
"""
import json, os, re, io
from datetime import datetime
from collections import defaultdict

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# ── Autenticación Google Drive ────────────────────────────────────────────────
SA_JSON = os.environ["GOOGLE_SERVICE_ACCOUNT"]
creds   = service_account.Credentials.from_service_account_info(
    json.loads(SA_JSON),
    scopes=["https://www.googleapis.com/auth/drive.readonly"]
)
drive = build("drive", "v3", credentials=creds, cache_discovery=False)

def download_xlsx(file_id):
    """Descarga un Google Sheet como XLSX y retorna un BytesIO."""
    req = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    buf = io.BytesIO()
    dl  = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf

# ── IDs de fuentes ────────────────────────────────────────────────────────────
IDS = {
    "pedidos":      "1IUJGqOfZHQpkelWiIIQLC5wFkVtPsJPpUUfnSlif1YY",
    "manifiesto":   "1ClTJj92BuAnox15lPTkFiolSM99DqQJIwNVGRb6Y2D8",
    "retrasos":     "1FIoDT9gnnuV9VKeB0VJ4krwyUwN6btXnBOQ0dRgdS1c",
    "dts":          "1k0f-1cSYFuze7sBhcwBo9YDMLewoMVchjrKOiQnWx9M",
    "devoluciones": "1Vma7zqpBBoH9Guttoa_La7W88_3e7JaNL9WnEoT_MUI",
    "storage":      "1sLzUbA_oLkw-UC0Sp5vmGegW8yqBeMdxHcpDkv41Eyg",
    "cc":           "11awsNaM5feyLVc__FN9b5Eb7Lfmw05HgB97pTt9bPxA",
    "fulfillment":  "1i25UAypuip_zGUg45k5NILfNGW_fK1oZlDQ7nbZwq_g",
}

# ── Helpers ───────────────────────────────────────────────────────────────────
MESES   = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
           7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
CUTOFF  = datetime(2026, 3, 1)
CC_CUTOFF = datetime(2026, 6, 9)

def pa(s):
    try:
        if s and "-" in str(s) and len(str(s)) == 10:
            return datetime.strptime(str(s), "%Y-%m-%d")
        p = str(s).strip().split("/")
        return datetime(int(p[2]), int(p[1]), int(p[0]))
    except:
        return None

def minfo(d):
    return d.strftime("%Y-%m"), f"{MESES[d.month]} {d.year}", d.isocalendar()[1], f"Semana {d.isocalendar()[1]}"

def fmt(v):
    if isinstance(v, datetime) and v.year > 1970:
        return f"{v.day}/{v.month}/{v.year}"
    return ""

def to_int(v):
    try: return int(v or 0)
    except: return 0

def extract_var(html, v):
    m = re.search(rf"const {v}\s*=\s*(\[.*?\]);", html, re.DOTALL)
    if not m: return []
    return json.loads(re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", " ", m.group(1)), strict=False)

def replace_var(html, v, data):
    nj = json.dumps(data, ensure_ascii=False)
    return re.sub(rf"((?:const|var)\s+{v}\s*=\s*)\[.*?\];", rf"\g<1>{nj};", html, flags=re.DOTALL)

def replace_cc(html, data):
    nj = json.dumps(data, ensure_ascii=False)
    result, n = re.subn(
        r"(var CC_DATA\s*=\s*window\._ccAllRows\s*=\s*)\[.*?\];",
        rf"\g<1>{nj};", html, flags=re.DOTALL
    )
    if n == 0: print("WARNING: CC_DATA pattern not found")
    return result

# ── Cargar dashboard ──────────────────────────────────────────────────────────
DASH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "index.html")
with open(DASH, encoding="utf-8") as f:
    html = f.read()

canal_data  = extract_var(html, "CANAL_DATA");  fechas_canal = {r["fecha"] for r in canal_data}
all_data    = extract_var(html, "ALL_DATA");    fechas_all   = {r["fecha"] for r in all_data}
transp_data = extract_var(html, "TRANSP_DATA")
otd_daily   = extract_var(html, "OTD_DAILY")
ots_detail  = extract_var(html, "OTS_DETAIL");  fechas_ots   = {r["fecha"] for r in ots_detail}
dts_data    = extract_var(html, "DTS_DATA");    itins_dts    = {r["itinerario"] for r in dts_data}
detalle     = extract_var(html, "DETALLE");     itins_det    = {r["itinerario"] for r in detalle}

tr_key      = {(r["fecha"], r["transportista"]): r for r in transp_data}
otd_iso_map = {}
for r in otd_daily:
    d = pa(r.get("fecha", ""))
    if d: otd_iso_map[d.strftime("%Y-%m-%d")] = r

print("Descargando 8 fuentes desde Google Drive...")

# ── 1. MANIFIESTO ─────────────────────────────────────────────────────────────
print("  → Manifiesto")
wb = openpyxl.load_workbook(download_xlsx(IDS["manifiesto"]), read_only=True, data_only=True)
manif = defaultdict(lambda: defaultdict(int))
_mani_sheets = [s for s in wb.sheetnames if s in ("Base de datos", "HISTORICO")]
print(f"    Hojas Manifiesto: {_mani_sheets}")
for _sheet_name in _mani_sheets:
    ws = wb[_sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tk, fe_raw, tr = (row[3] if len(row)>3 else None), (row[2] if len(row)>2 else None), (row[4] if len(row)>4 else None)
        if not tk or not fe_raw: continue
        fecha = fe_raw if isinstance(fe_raw, datetime) else pa(str(fe_raw))
        if not fecha or tr not in ("Despachalo","Recibelo") or fecha < CUTOFF: continue
        manif[fecha.strftime("%Y-%m-%d")][tr] += 1

for fi, trs in manif.items():
    d = datetime.strptime(fi, "%Y-%m-%d"); mn,ms,sn,ss = minfo(d)
    for tr, cnt in trs.items():
        k = (fi, tr)
        if k in tr_key: tr_key[k]["count"] = cnt
        else: tr_key[k] = {"fecha":fi,"dia":f"{d.day}/{d.month}/{d.year}","mes_num":mn,
                            "mes":ms,"semana":ss,"sem_num":sn,"transportista":tr,"count":cnt}
transp_data = sorted(tr_key.values(), key=lambda x: (x["fecha"], x["transportista"]))
daily_man   = {fi: sum(v.values()) for fi,v in manif.items()}

# ── 2. RETRASOS MELI ──────────────────────────────────────────────────────────
print("  → Retrasos MELI")
wb = openpyxl.load_workbook(download_xlsx(IDS["retrasos"]), read_only=True, data_only=True)
ws = wb["Categorizacion"]
hdr = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
ots_i=hdr.index("OTS"); otd_i=hdr.index("OTD")
cat_i=hdr.index("Categoria"); courr_i=hdr.index("Courrier")
coment_i=hdr.index("Comentario SP"); com_i=hdr.index("Comuna")

otd_errors=defaultdict(int); ots_errors=defaultdict(int)
otd_det_new=[]; ots_det_new=[]; delays_agg=defaultdict(int)

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    d = row[0] if isinstance(row[0], datetime) else None
    if not d:
        try: p=str(row[0]).split("/"); d=datetime(int(p[2]),int(p[1]),int(p[0]))
        except: continue
    if d < CUTOFF: continue
    fi=d.strftime("%Y-%m-%d"); dia=f"{d.day}/{d.month}/{d.year}"; mn,ms,sn,ss=minfo(d)
    ots_v=to_int(row[ots_i]); otd_v=to_int(row[otd_i])
    venta=str(row[1]).strip() if row[1] else ""
    cat=str(row[cat_i]).strip() if row[cat_i] else ""
    courr=str(row[courr_i]).strip() if row[courr_i] else ""
    coment=str(row[coment_i]).strip() if row[coment_i] else ""
    comuna=str(row[com_i]).strip() if row[com_i] else ""
    if ots_v==1:
        ots_errors[fi]+=1
        if fi not in fechas_ots:
            ots_det_new.append({"fecha":fi,"dia":dia,"venta":venta,"categoria":cat,
                "comentario_sp":coment,"mes_num":mn,"mes":ms,"semana":ss,"sem_num":sn})
    if otd_v==1:
        otd_errors[fi]+=1; delays_agg[(courr,sn,mn,ms)]+=1
        otd_det_new.append({"fecha":dia,"venta":venta,"categoria":cat,
            "courrier":courr,"comentario_sp":coment,"comuna":comuna,
            "mes_num":mn,"mes":ms,"sem_num":sn,"semana":ss})

for fi, tot in daily_man.items():
    d=datetime.strptime(fi,"%Y-%m-%d"); mn,ms,sn,ss=minfo(d)
    dia=f"{d.day}/{d.month}/{d.year}"; err=otd_errors.get(fi,0)
    rate=round((tot-err)/tot*100,1) if tot>0 else 100.0
    if fi in otd_iso_map:
        otd_iso_map[fi].update({"pedidos":tot,"errores_otd":err,"otd_rate":rate,
                                 "dia":dia,"fecha":dia,"mes_num":mn,"mes":ms,"sem_num":sn,"semana":ss})
    else:
        otd_iso_map[fi]={"fecha":dia,"dia":dia,"pedidos":tot,"errores_otd":err,
                         "otd_rate":rate,"mes_num":mn,"mes":ms,"semana":ss,"sem_num":sn}

otd_daily  = [otd_iso_map[k] for k in sorted(otd_iso_map.keys())]
ots_detail.extend(ots_det_new)
otd_detail = otd_det_new
delays_tr  = [{"courrier":c,"sem_num":sn,"mes_num":mn,"mes":ms,"delays":cnt}
              for (c,sn,mn,ms),cnt in sorted(delays_agg.items())]

# ── 3. PEDIDOS x MARKETPLACE ──────────────────────────────────────────────────
print("  → Pedidos x Marketplace")
wb_ped = openpyxl.load_workbook(download_xlsx(IDS["pedidos"]), read_only=True, data_only=True)
ws_ped = wb_ped["Pedidos"]
source_canal = defaultdict(dict)
for row in ws_ped.iter_rows(min_row=2, values_only=True):
    if not row[0] or not row[1]: continue
    fecha = row[0] if isinstance(row[0], datetime) else pa(str(row[0]))
    if not fecha or fecha < CUTOFF: continue
    cn=str(row[1]).strip(); fi=fecha.strftime("%Y-%m-%d")
    if cn: source_canal[fi][cn]={"pedidos":int(row[2] or 0),"entregados":int(row[3] or 0)}

for r in canal_data:
    src=source_canal.get(r["fecha"],{}).get(r["canal"])
    if src and (r["entregados"]!=src["entregados"] or r["pedidos"]!=src["pedidos"]):
        r["entregados"]=src["entregados"]; r["pedidos"]=src["pedidos"]
        r["nivel"]=round(src["entregados"]/src["pedidos"]*100,1) if src["pedidos"]>0 else 0.0

new_canal=[]
for fi, canales in sorted(source_canal.items()):
    if fi not in fechas_canal:
        d=datetime.strptime(fi,"%Y-%m-%d"); mn,ms,sn,ss=minfo(d); dia=f"{d.day}/{d.month}/{d.year}"
        for cn,v in canales.items():
            new_canal.append({"fecha":fi,"dia":dia,"canal":cn,"pedidos":v["pedidos"],
                "entregados":v["entregados"],
                "nivel":round(v["entregados"]/v["pedidos"]*100,1) if v["pedidos"]>0 else 0.0,
                "mes_num":mn,"mes":ms,"semana":ss,"sem_num":sn})
canal_data.extend(new_canal); canal_data.sort(key=lambda x:(x["fecha"],x["canal"]))

for fi, canales in source_canal.items():
    tp=sum(v["pedidos"] for v in canales.values())
    te=sum(v["entregados"] for v in canales.values())
    err=ots_errors.get(fi,0)
    rate=round((tp-err)/tp*100,1) if tp>0 else 0.0
    d=datetime.strptime(fi,"%Y-%m-%d"); mn,ms,sn,ss=minfo(d); dia=f"{d.day}/{d.month}/{d.year}"
    entry={"fecha":fi,"pedidos":tp,"errores":err,"entregados":tp-err,
           "ots_rate":rate,"dia":dia,"mes_num":mn,"mes":ms,"semana":ss,"sem_num":sn}
    if fi in fechas_all:
        for r in all_data:
            if r["fecha"]==fi: r.update(entry); break
    else: all_data.append(entry)
all_data.sort(key=lambda x: x["fecha"])

# ── 4. DTS ────────────────────────────────────────────────────────────────────
print("  → DTS")
wb=openpyxl.load_workbook(download_xlsx(IDS["dts"]), read_only=True, data_only=True)
ws_emb=wb["Embarques"]; ws_det_sh=wb["Detalle"]

for row in ws_emb.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    itin=str(row[0]).strip()
    sr=str(row[2]).lower().strip() if len(row)>2 and row[2] else ""
    status=("Finalizado" if "finaliz" in sr else "Transito" if "transito" in sr or "tránsito" in sr
            else "Booking" if "booking" in sr else "Asignado" if "asig" in sr else "Pendiente")
    eta_r=fmt(row[8] if len(row)>8 else None)
    fb=fmt(row[40] if len(row)>40 else None)
    fc=fmt(row[42] if len(row)>42 else None)
    eta_d=pa(eta_r)
    if eta_d and eta_d<CUTOFF: continue
    ref=(row[40] if isinstance(row[40] if len(row)>40 else None,datetime) and row[40].year>1970
         else row[8] if isinstance(row[8] if len(row)>8 else None,datetime) and row[8].year>1970
         else datetime(2026,6,1))
    if not isinstance(ref,datetime): ref=datetime(2026,6,1)
    mn,ms,sn,ss=minfo(ref)
    bod=pa(fb); chk=pa(fc)
    if bod and chk and chk.year>1900:
        diff=(chk-bod).days; dts_lbl="100%" if diff<=1 else "0%"; dts_pct=100.0 if diff<=1 else 0.0
    else: dts_lbl="Pendiente"; dts_pct=None; diff=None
    entry={"itinerario":itin,"status":status,"eta_real":eta_r,"fecha_bodega":fb,
           "fecha_check":fc or "1/1/1900","dts":dts_lbl,"dts_pct":dts_pct,
           "dts_dias":diff,"mes_num":mn,"mes":ms,"sem_num":sn,"semana":ss}
    if itin in itins_dts:
        for r in dts_data:
            if r["itinerario"]==itin: r.update(entry); break
    else: dts_data.append(entry)

valid_itins={r["itinerario"] for r in dts_data}
for row in ws_det_sh.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    itin=str(row[0]).strip()
    if itin not in valid_itins or itin in itins_det: continue
    sku_cl=str(row[2]).strip() if row[2] else ""
    if not sku_cl: continue
    detalle.append({"itinerario":itin,"sku_cn":str(row[1]).strip() if row[1] else "",
        "sku_cl":sku_cl,"etd":fmt(row[3]),"eta":fmt(row[4]),"fecha_bodega":fmt(row[5]),
        "descripcion":str(row[6]).strip() if row[6] else "",
        "marca":str(row[7]).strip() if row[7] else "",
        "clasificacion":str(row[8]).strip() if row[8] else "",
        "proveedor":str(row[9]).strip() if row[9] else "",
        "puerto":str(row[10]).strip() if len(row)>10 and row[10] else "",
        "cantidad":str(int(row[11])) if len(row)>11 and isinstance(row[11],(int,float))
                   else (str(row[11]).strip() if len(row)>11 and row[11] else "")})

# ── 5. DEVOLUCIONES ───────────────────────────────────────────────────────────
print("  → Devoluciones")
wb=openpyxl.load_workbook(download_xlsx(IDS["devoluciones"]), read_only=True, data_only=True)
# Buscar la hoja correcta (puede llamarse Dev.2026, Devoluciones, Sheet1, etc.)
_dev_candidates = [s for s in wb.sheetnames if "dev" in s.lower() or "devol" in s.lower()]
_dev_sheet = _dev_candidates[0] if _dev_candidates else wb.sheetnames[0]
print(f"    Hojas disponibles: {wb.sheetnames} → usando '{_dev_sheet}'")
ws=wb[_dev_sheet]
# Mapear columnas por nombre de encabezado (robusto a cambios de orden/estructura)
_dev_hdr=[str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
_dev_idx={h:i for i,h in enumerate(_dev_hdr)}
print(f"    Encabezados Devoluciones: {_dev_hdr}")
i_fecha=_dev_idx.get("Fecha",1)
i_mes=_dev_idx.get("MES",2)
i_sku=_dev_idx.get("SKU",3)
i_canal=_dev_idx.get("MarketPlace",4)
i_espec=_dev_idx.get("Especificacion",5)
i_clasi=_dev_idx.get("Clasificacion",_dev_idx.get("Clasificación",8))
i_sem=_dev_idx.get("Semana",11)
dev_rows=[]
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[i_fecha]: continue
    fecha=row[i_fecha] if isinstance(row[i_fecha],datetime) else pa(str(row[i_fecha]))
    if not fecha: continue
    fi=fecha.strftime("%Y-%m-%d")
    mes=str(row[i_mes]).strip() if row[i_mes] else f"{MESES[fecha.month]} {fecha.year}"
    canal=str(row[i_canal]).strip() if row[i_canal] else ""
    espec=str(row[i_espec]).strip() if row[i_espec] else ""
    clasi=str(row[i_clasi]).strip() if row[i_clasi] else "DESCONOCIDA"
    sku=str(row[i_sku]).strip() if row[i_sku] else ""
    raw_sem=str(row[i_sem]).strip() if len(row)>i_sem and row[i_sem] else ""
    if raw_sem:
        core=raw_sem.lstrip("Ww")
        sem=f"W{int(float(core))}" if core.replace(".","").isdigit() else f"W{fecha.isocalendar()[1]}"
    else: sem=f"W{fecha.isocalendar()[1]}"
    dev_rows.append([fi,mes,canal,espec,clasi,sem,sku])
dev_rows.sort(key=lambda r: r[0])

# ── 6. STORAGE ────────────────────────────────────────────────────────────────
print("  → Storage")
wb_st=openpyxl.load_workbook(download_xlsx(IDS["storage"]), read_only=True, data_only=True)
ws=wb_st["Ocupacion"]
ocup_fechas=[]; ocup_pcts=[]; ocup_occ=[]; ocup_disp=[]
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    fecha=row[0] if isinstance(row[0],datetime) else pa(str(row[0]))
    if not fecha: continue
    disp=int(row[1] or 0); ocup=int(row[2] or 0); total=disp+ocup
    ocup_fechas.append(f"{fecha.day}/{fecha.month}/{fecha.year}")
    ocup_pcts.append(round(ocup/total*100,2) if total>0 else 0.0)
    ocup_occ.append(ocup); ocup_disp.append(disp)

ws_stk=wb_st["Stock"]
stk_fisico=stk_disp_v=0
for row in ws_stk.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    stk_fisico+=int(row[9] or 0); stk_disp_v+=int(row[11] or 0)

ws_mae=wb_st["Maestro de ubicaciones"]
storage_total_ub=sum(1 for row in ws_mae.iter_rows(min_row=2, values_only=True) if row[0])

# ── 7. CONTEOS CÍCLICOS ───────────────────────────────────────────────────────
print("  → Conteos Cíclicos")
wb=openpyxl.load_workbook(download_xlsx(IDS["cc"]), read_only=True, data_only=True)
ws=wb["Detalle diario"]
hdr_cc=[str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
idx={h:i for i,h in enumerate(hdr_cc)}
contado_i=idx.get("✓ Contado",12); stk_teo_i=idx.get("Stock teórico",13)
stk_app_i=idx.get("Stock App ET Brands",14); stk_cnt_i=idx.get("Stock contado",idx.get("Stock Físico",15))
ira_i=idx.get("IRA",19); desc_i=idx.get("Comentario",18)
cc_data=[]
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    if row[contado_i] is not True: continue
    fecha=row[0] if isinstance(row[0],datetime) else pa(str(row[0]))
    if not fecha or fecha<CC_CUTOFF: continue
    familia=str(row[5]).strip() if row[5] else ""
    sku=str(row[6]).strip() if row[6] else ""
    abc=str(row[8]).strip() if row[8] else ""
    stk_teo=int(row[stk_teo_i]) if row[stk_teo_i] is not None else None
    try: stk_app=int(row[stk_app_i]) if row[stk_app_i] is not None else None
    except: stk_app=None
    stk_cnt=int(row[stk_cnt_i]) if row[stk_cnt_i] is not None else None
    diff=int(row[16]) if len(row)>16 and row[16] is not None else 0
    ira=round(float(row[ira_i]),10) if row[ira_i] is not None else None
    desc=str(row[desc_i]).strip() if row[desc_i] else ""
    cc_data.append({"fecha":fecha.strftime("%Y-%m-%d"),"dia":f"{fecha.day}/{fecha.month}/{fecha.year}",
        "sku":sku,"mp":familia,"marca":familia,"abc":abc,
        "stk_teo":stk_teo,"stk_app":stk_app,"stk_cnt":stk_cnt,"diff":diff,"ira":ira,
        "teo":stk_teo,"app":stk_app,"cnt":stk_cnt,"desc":desc})
cc_data.sort(key=lambda r: r["fecha"])

# ── 8. CONTROL FULFILLMENT ────────────────────────────────────────────────────
print("  → Fulfillment")
wb_ff=openpyxl.load_workbook(download_xlsx(IDS["fulfillment"]), read_only=True, data_only=True)
ws_ff=wb_ff["Control"]
ff_detail=[]
for row in ws_ff.iter_rows(min_row=6, values_only=True):
    if not row or row[1] is None: continue
    full_n=str(int(row[1])) if isinstance(row[1],float) else str(row[1]).strip()
    mp=str(row[2]).strip().upper() if row[2] else ""
    f_car=fmt(row[3]); f_ent=fmt(row[4])
    sol=int(row[8]) if row[8] is not None else 0
    ent=int(row[9]) if row[9] is not None else None
    estado=str(row[10]).strip() if row[10] else ""
    coment=str(row[12]).strip() if row[12] else ""
    if not estado:
        chk=row[5] is True; recl=row[6] is True; etiq=row[7] is True
        if chk and recl and etiq: estado="Procesamiento finalizado"
        elif chk and recl: estado="Etiquetado pendiente"
        elif chk: estado="Recolección pendiente"
        else: estado="Pendiente ingreso"
    ff_detail.append([full_n,mp,f_car,f_ent,sol,ent,estado,coment])

# ── Filtro cutoff ─────────────────────────────────────────────────────────────
canal_data  = [r for r in canal_data  if (d:=pa(r.get("fecha",""))) and d>=CUTOFF]
all_data    = [r for r in all_data    if (d:=pa(r.get("fecha",""))) and d>=CUTOFF]
transp_data = [r for r in transp_data if (d:=pa(r.get("fecha",""))) and d>=CUTOFF]
ots_detail  = [r for r in ots_detail  if (d:=pa(r.get("fecha",""))) and d>=CUTOFF]
delays_tr   = [r for r in delays_tr   if r.get("mes_num","")>="2026-03"]
dts_data    = [r for r in dts_data    if not r.get("eta_real") or ((d:=pa(r["eta_real"])) and d>=CUTOFF)]
valid2      = {r["itinerario"] for r in dts_data}
detalle     = [r for r in detalle     if r.get("itinerario") in valid2]

# ── Guardar ───────────────────────────────────────────────────────────────────
print("Actualizando index.html...")
for var, data in [
    ("CANAL_DATA",canal_data),("ALL_DATA",all_data),("TRANSP_DATA",transp_data),
    ("OTD_DAILY",otd_daily),("OTS_DETAIL",ots_detail),("OTD_DETAIL",otd_detail),
    ("DELAYS_TRANSP",delays_tr),("DTS_DATA",dts_data),("DETALLE",detalle),
    ("DEV_ROWS",dev_rows),("OCUP_FECHAS",ocup_fechas),("OCUP_PCTS",ocup_pcts),
    ("OCUP_OCC",ocup_occ),("OCUP_DISP",ocup_disp),("FF_DETAIL",ff_detail),
]:
    html = replace_var(html, var, data)

for pat, val in [
    (r"const STOCK_FISICO\s*=\s*\d+;",     f"const STOCK_FISICO={stk_fisico};"),
    (r"const STOCK_DISP_TOT\s*=\s*\d+;",   f"const STOCK_DISP_TOT={stk_disp_v};"),
    (r"const STORAGE_TOTAL_UB\s*=\s*\d+;", f"const STORAGE_TOTAL_UB={storage_total_ub};"),
]:
    html = re.sub(pat, val, html)

html = replace_cc(html, cc_data)
hoy  = datetime.now()
html = re.sub(r"Datos al \d+/\d+/\d+", f"Datos al {hoy.day}/{hoy.month}/{hoy.year}", html)

with open(DASH, "w", encoding="utf-8") as f:
    f.write(html)

print(f"✅ Dashboard actualizado — {len(html)//1024} KB")
print(f"   Fecha: {hoy.day}/{hoy.month}/{hoy.year}")
print(f"   CANAL_DATA: {len(canal_data)} regs hasta {max(r['fecha'] for r in canal_data)}")
print(f"   OTS_DETAIL: {len(ots_detail)} regs")
print(f"   DTS_DATA:   {len(dts_data)} embarques")
print(f"   CC_DATA:    {len(cc_data)} regs")
